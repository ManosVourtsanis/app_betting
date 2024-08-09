import openpyxl
import re
import xlwings as xw
import subprocess


# Load the workbook and select the active worksheet
filename = 'Bets By Tolis_v2.xlsx'

def load_workbook():
    return openpyxl.load_workbook(filename)

"""def force_recalculation(filename):
    app = xw.App(visible=False)  # Run Excel in the background
    wb = app.books.open(filename)
    wb.app.calculate()  # Forces a recalculation of the entire workbook
    wb.save()
    wb.close()
    app.quit()"""

# Function to find the first empty row in a specific column (e.g., Column A)
def find_first_empty_row_in_column(sheet, column):
    for row in range(2, sheet.max_row + 1):  # Starting from 2 assuming row 1 has headers
        if sheet[f'{column}{row}'].value is None:
            return row
    return sheet.max_row + 1  # Return next row if no empty cell is found

# Function to get a valid integer input
def get_valid_integer(prompt):
    while True:
        try:
            value = int(input(prompt))
            return value
        except ValueError:
            print("Please enter a valid number.")

# Function to validate float input with comma handling
def get_valid_float(prompt):
    while True:
        value = input(prompt).replace(',', '.')
        if not re.match(r"^\d+(\.\d+)?$", value):
            print("Please enter a valid number.")
        else:
            return float(value)

# Function to calculate total odds based on cell values
def calculate_total_odds(cells):
    total_odds = 1
    for cell in cells:
        value = cell if cell else 1
        try:
            total_odds *= float(value)
        except ValueError:
            total_odds *= 1
    return round(total_odds, 3)

def insert_data():
    workbook = load_workbook()
    sheet = workbook.active
    
    data = []
    num_rows = get_valid_integer("How many rows of data do you want to add? ")

    for _ in range(num_rows):
        date = input("Enter Date (DD-MM-YYYY): ")
        
        num_matches = get_valid_integer("How many matches do you want to add (1-5)? ")
        while num_matches < 1 or num_matches > 5:
            num_matches = get_valid_integer("How many matches do you want to add (1-5)? ")
        matches = []
        odds = []

        for i in range(num_matches):
            match = input(f"Enter Match No{i+1}: ")
            odd = get_valid_float(f"Enter Odd_{i+1}: ")  # Read as string for later validation
            matches.append(match)
            odds.append(odd)

        # Append empty strings for matches and odds if fewer than 5 matches
        while len(matches) < 5:
            matches.append("")
        while len(odds) < 5:
            odds.append("")

        stake = float(input("Enter Stake: "))

        # Calculate the total odds based on the provided formula
        total_odds = calculate_total_odds(odds)

        # Default result is "Pending"
        result = "Pending"

        # Append row data to the data list
        data.append([
            date,
            matches[0], odds[0],
            matches[1], odds[1],
            matches[2], odds[2],
            matches[3], odds[3],
            matches[4], odds[4],
            stake, total_odds, result
        ])

    # Find the first empty row in the "Date" column (Column A)
    first_empty_row = find_first_empty_row_in_column(sheet, 'A')

    # Write the data to the sheet starting from the first empty row
    for i, row in enumerate(data, start=first_empty_row):
        for j, value in enumerate(row, start=1):
            if j not in [13, 15, 16]:  # Avoid writing to columns with formulas
                sheet.cell(row=i, column=j, value=value)


    # Save the workbook
    workbook.save('Bets By Tolis_v2.xlsx')
    #force_recalculation(filename)
    print("Data successfully inserted.")

def format_value(value):
    return str(value) if value is not None else ""

def view_data():

    workbook = openpyxl.load_workbook('Bets By Tolis_v2.xlsx', data_only=True)
    sheet = workbook.active

    print("Viewing data:")
    date_column = 'A'
    max_row = 500
    print(f"Total number of rows in the sheet: {max_row}")
    

    # Print header
    print(f"{'Row':<5} {'Match 01':<15} {'Odd 01':<10} {'Match 02':<15} {'Odd 02':<10} {'Match 03':<15} {'Odd 03':<10} {'Match 04':<15} {'Odd 04':<10} {'Match 05':<15} {'Odd 05':<10} {'Stake':<10} {'Total Odds':<12} {'Result':<10} {'Profit/Lose':<12} {'Units':<15}")

    for row in range(2, max_row + 1):  # Starting from 2 assuming row 1 has headers
        date_value = sheet[f'{date_column}{row}'].value
        if date_value is not None:
            row_number = row - 1 
            matches = [
                format_value(sheet.cell(row=row, column=2).value),  # Match No1
                format_value(sheet.cell(row=row, column=4).value),  # Match No2
                format_value(sheet.cell(row=row, column=6).value),  # Match No3
                format_value(sheet.cell(row=row, column=8).value),  # Match No4
                format_value(sheet.cell(row=row, column=10).value)  # Match No5
            ]
            odds = [
                format_value(sheet.cell(row=row, column=3).value),  # Odd_01
                format_value(sheet.cell(row=row, column=5).value),  # Odd_02
                format_value(sheet.cell(row=row, column=7).value),  # Odd_03
                format_value(sheet.cell(row=row, column=9).value),  # Odd_04
                format_value(sheet.cell(row=row, column=11).value)  # Odd_05
            ]
            stake = format_value(sheet.cell(row=row, column=12).value)
            #total_odds = calculate_total_odds(odds)
            total_odds = format_value(sheet.cell(row=row, column=13).value)
            result = format_value(sheet.cell(row=row, column=14).value)
            profit_lose = format_value(sheet.cell(row=row, column=15).value)
            units = format_value(sheet.cell(row=row, column=16).value)

            # Print row data
            print(f"{row_number+1:<5} {matches[0]:<15} {odds[0]:<10} {matches[1]:<15} {odds[1]:<10} {matches[2]:<15} {odds[2]:<10} {matches[3]:<15} {odds[3]:<10} {matches[4]:<15} {odds[4]:<10} {stake:<10} {total_odds:<12} {result:<10} {profit_lose:<12} {units:<15}")

def delete_rows():
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    # Get the total number of rows in the sheet
    max_row = 500
    
    # Prompt user for the row number to clear
    row_to_clear = get_valid_integer(f"Enter the row number you want to clear (2 to {max_row}): ")

    # Validate that the row number is within the acceptable range
    if row_to_clear < 2 or row_to_clear > max_row:
        print(f"Invalid row number. Please enter a number between 2 and {max_row}.")
        return

    # Get the total number of columns in the sheet
    max_col = sheet.max_column
    # Define columns to ignore (e.g., columns with formulas)
    ignore_columns = [13, 15, 16]  # Columns C, E, and F are examples; adjust as needed
    
    # Clear the contents of the specified row, ignoring specified columns
    for col in range(1, max_col + 1):
        if col not in ignore_columns:
            cell = sheet.cell(row=row_to_clear, column=col)
            cell.value = None
    
    # Shift up the rows below
    for row in range(row_to_clear + 1, max_row + 1):
        for col in range(1, max_col + 1):
            if col not in ignore_columns:
                cell_above = sheet.cell(row=row - 1, column=col)
                cell_below = sheet.cell(row=row, column=col)
                cell_above.value = cell_below.value
                cell_below.value = None
    
    # Save the workbook
    workbook.save(filename)
    #force_recalculation(filename)
    print(f"Row {row_to_clear} successfully cleared.")

def menu():
    while True:
        print("\nMenu:")
        print("1. Insert Data")
        print("2. View Data")
        print("3. Delete Rows")
        print("4. Exit")

        choice = get_valid_integer("Enter your choice (1-4): ")

        if choice == 1:
            insert_data()
        elif choice == 2:
            view_data()
        elif choice == 3:
            delete_rows()
        elif choice == 4:
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 4.")

if __name__ == "__main__":
    menu()
